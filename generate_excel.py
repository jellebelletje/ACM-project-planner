#!/usr/bin/env python3
"""Generate an Excel workbook with all 6 sheets pre-populated with seed data.
Import this directly into Google Sheets."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_workbook():
    with open('seed-data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    wrap_align = Alignment(wrap_text=True, vertical='top')
    top_align = Alignment(vertical='top')
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0')
    )

    # ---- Sheet 1: Activities ----
    ws = wb.active
    ws.title = 'Activities'
    headers = ['id', 'title', 'intro_text', 'full_description', 'pdca_phase',
               'sequence', 'status', 'due_date', 'depends_on',
               'particularisation_guidance', 'created_at', 'updated_at']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Status dropdown validation
    status_dv = DataValidation(type='list', formula1='"not_started,in_progress,completed,blocked"')
    status_dv.error = 'Invalid status'
    status_dv.errorTitle = 'Invalid Status'
    ws.add_data_validation(status_dv)

    for i, act in enumerate(data['activities'], 2):
        ws.cell(row=i, column=1, value=act.get('id', ''))
        ws.cell(row=i, column=2, value=act.get('title', ''))
        ws.cell(row=i, column=3, value=act.get('intro_text', '')).alignment = wrap_align
        ws.cell(row=i, column=4, value=act.get('full_description', '')).alignment = wrap_align
        ws.cell(row=i, column=5, value=act.get('pdca_phase', ''))
        ws.cell(row=i, column=6, value=act.get('sequence', ''))
        status_cell = ws.cell(row=i, column=7, value=act.get('status', 'not_started'))
        ws.cell(row=i, column=8, value=act.get('due_date', ''))
        ws.cell(row=i, column=9, value=act.get('depends_on', ''))
        ws.cell(row=i, column=10, value=act.get('particularisation_guidance', '')).alignment = wrap_align
        ws.cell(row=i, column=11, value='')  # created_at
        ws.cell(row=i, column=12, value='')  # updated_at
        status_dv.add(status_cell)

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 60
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.freeze_panes = 'A2'

    # ---- Sheet 2: Todos ----
    ws2 = wb.create_sheet('Todos')
    headers2 = ['id', 'activity_id', 'text', 'is_done', 'is_project_specific',
                'assigned_to', 'due_date', 'sequence']

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, todo in enumerate(data['todos'], 2):
        ws2.cell(row=i, column=1, value=todo.get('id', ''))
        ws2.cell(row=i, column=2, value=todo.get('activity_id', ''))
        ws2.cell(row=i, column=3, value=todo.get('text', '')).alignment = wrap_align
        ws2.cell(row=i, column=4, value=bool(todo.get('is_done', False)))
        ws2.cell(row=i, column=5, value=bool(todo.get('is_project_specific', False)))
        ws2.cell(row=i, column=6, value=todo.get('assigned_to', ''))
        ws2.cell(row=i, column=7, value=todo.get('due_date', ''))
        ws2.cell(row=i, column=8, value=todo.get('sequence', 0))

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 80
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 10
    ws2.freeze_panes = 'A2'

    # ---- Sheet 3: Questions ----
    ws3 = wb.create_sheet('Questions')
    headers3 = ['id', 'activity_id', 'sub_topic', 'question_text', 'ask_whom',
                'answer', 'is_answered', 'sequence']

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, q in enumerate(data['questions'], 2):
        ws3.cell(row=i, column=1, value=q.get('id', ''))
        ws3.cell(row=i, column=2, value=q.get('activity_id', ''))
        ws3.cell(row=i, column=3, value=q.get('sub_topic', ''))
        ws3.cell(row=i, column=4, value=q.get('question_text', '')).alignment = wrap_align
        ws3.cell(row=i, column=5, value=q.get('ask_whom', ''))
        ws3.cell(row=i, column=6, value=q.get('answer', '')).alignment = wrap_align
        ws3.cell(row=i, column=7, value=bool(q.get('is_answered', False)))
        ws3.cell(row=i, column=8, value=q.get('sequence', 0))

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 80
    ws3.column_dimensions['E'].width = 35
    ws3.column_dimensions['F'].width = 50
    ws3.column_dimensions['G'].width = 14
    ws3.column_dimensions['H'].width = 10
    ws3.freeze_panes = 'A2'

    # ---- Sheet 4: Notes_Links ----
    ws4 = wb.create_sheet('Notes_Links')
    headers4 = ['id', 'activity_id', 'type', 'content', 'url', 'label', 'date_added']

    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    type_dv = DataValidation(type='list', formula1='"note,link,attachment_ref"')
    ws4.add_data_validation(type_dv)

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 14
    ws4.column_dimensions['D'].width = 50
    ws4.column_dimensions['E'].width = 40
    ws4.column_dimensions['F'].width = 25
    ws4.column_dimensions['G'].width = 18
    ws4.freeze_panes = 'A2'

    # ---- Sheet 5: Technical_Milestones ----
    ws5 = wb.create_sheet('Technical_Milestones')
    headers5 = ['id', 'milestone_name', 'date', 'status', 'notes', 'sequence', 'timeline_type']

    for col, h in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    ms_status_dv = DataValidation(type='list', formula1='"planned,in_progress,completed,delayed"')
    ws5.add_data_validation(ms_status_dv)

    ms_timeline_dv = DataValidation(type='list', formula1='"technical,acm"')
    ws5.add_data_validation(ms_timeline_dv)

    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 30
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 14
    ws5.column_dimensions['E'].width = 40
    ws5.column_dimensions['F'].width = 10
    ws5.column_dimensions['G'].width = 16
    ws5.freeze_panes = 'A2'

    # ---- Sheet 6: Project_Config ----
    ws6 = wb.create_sheet('Project_Config')
    headers6 = ['key', 'value']

    for col, h in enumerate(headers6, 1):
        cell = ws6.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    config_rows = [
        ('project_name', ''),
        ('client_name', ''),
        ('start_date', ''),
        ('end_date', ''),
        ('current_phase', 'Plan I: Diagnosis'),
        ('consultant_name', ''),
    ]
    for i, (key, value) in enumerate(config_rows, 2):
        ws6.cell(row=i, column=1, value=key)
        ws6.cell(row=i, column=2, value=value)

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 40
    ws6.freeze_panes = 'A2'

    # Save
    filename = 'ACM_Dashboard_Data.xlsx'
    wb.save(filename)
    print(f"Created {filename}")
    print(f"  Activities: {len(data['activities'])} rows")
    print(f"  Todos: {len(data['todos'])} rows")
    print(f"  Questions: {len(data['questions'])} rows")
    print(f"  Notes_Links: 0 rows (empty)")
    print(f"  Technical_Milestones: 0 rows (empty)")
    print(f"  Project_Config: {len(config_rows)} rows")
    print(f"\nImport into Google Sheets: File > Import > Upload > Replace spreadsheet")


if __name__ == '__main__':
    create_workbook()
